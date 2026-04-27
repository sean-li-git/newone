#!/usr/bin/env python3
"""生成 Offer审批助手 所需的 4 个 Excel 模板"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DIR = os.path.dirname(os.path.abspath(__file__))

HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='2563EB')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
CELL_FONT = Font(name='Arial', size=11)
THIN_BORDER = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0'),
)
HINT_FONT = Font(name='Arial', size=10, color='94A3B8', italic=True)
BLUE_FONT = Font(name='Arial', size=11, color='0000FF')
COMPUTED_FILL = PatternFill('solid', fgColor='F1F5F9')

def style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

def style_data_row(ws, row, max_col, computed_cols=None):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = CELL_FONT
        cell.border = THIN_BORDER
        if computed_cols and c in computed_cols:
            cell.fill = COMPUTED_FILL

# ========================
# 1. 薪酬模板 (salary-template.xlsx)
# ========================
def gen_salary_template():
    wb = Workbook()
    ws = wb.active
    ws.title = '薪酬信息'

    headers = [
        ('币种', 10), ('国家/城市', 14), ('职级', 10), ('职位通道', 12), ('职位类型', 12),
        ('基本月薪', 14), ('津贴及其他现金补贴', 18), ('绩效（月均）', 14), ('年终奖', 14),
        ('入职授予（年化）', 16), ('签字股票总额', 14), ('签字费总额', 14), ('安家费', 12),
        ('现金收入', 14), ('年收入', 14), ('年收入（含一次性）', 18),
    ]
    for i, (h, w) in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    style_header(ws, 1, len(headers))

    computed = {14, 15, 16}
    # 示例行
    sample = ['CNY', '中国-上海', 'L7', '技术', '研发', 50000, 2000, 5000, 80000, 100000, 200000, 100000, 30000, '', '', '']
    for i, v in enumerate(sample, 1):
        cell = ws.cell(row=2, column=i, value=v)
        cell.font = BLUE_FONT if i <= 13 and v != '' else CELL_FONT
        cell.border = THIN_BORDER
        if i in computed:
            cell.fill = COMPUTED_FILL
    # 公式
    ws.cell(row=2, column=14, value='=F2*12+H2*12+I2+G2')
    ws.cell(row=2, column=15, value='=N2+J2')
    ws.cell(row=2, column=16, value='=O2+K2/2+L2/2+M2/2')

    # 说明 sheet
    ws2 = wb.create_sheet('填写说明')
    notes = [
        ['字段', '说明', '示例'],
        ['币种', '薪酬币种代码', 'CNY / USD / HKD'],
        ['国家/城市', '工作地，格式：国家-城市', '中国-上海 / US-SF'],
        ['职级', '公司内部职级', 'L5 / L7 / M1'],
        ['职位通道', '职位所属通道', '技术 / 产品 / 设计'],
        ['职位类型', '职位族分类', '研发 / 运营 / 市场'],
        ['基本月薪', '每月固定底薪', '50000'],
        ['津贴及其他现金补贴', '月度/年度津贴补贴', '2000'],
        ['绩效（月均）', '绩效奖金月均值', '5000'],
        ['年终奖', '年度奖金（全额）', '80000'],
        ['入职授予（年化）', '股票/期权年化价值', '100000'],
        ['签字股票总额', '签字股票总额（一次性）', '200000'],
        ['签字费总额', '签字费总额（一次性）', '100000'],
        ['安家费', '搬迁/安家补贴（一次性）', '30000'],
        ['现金收入', '=基本月薪×12+绩效×12+年终奖+津贴 [系统自动计算]', ''],
        ['年收入', '=现金收入+入职授予年化 [系统自动计算]', ''],
        ['年收入（含一次性）', '=年收入+签字股票/N+签字费/N+安家费/N [系统自动计算]', ''],
    ]
    for r, row_data in enumerate(notes, 1):
        for c, v in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=v)
    style_header(ws2, 1, 3)
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 50
    ws2.column_dimensions['C'].width = 20

    wb.save(os.path.join(DIR, 'salary-template.xlsx'))
    print('✅ salary-template.xlsx')

# ========================
# 2. 历史 Offer 批量导入模板
# ========================
def gen_history_template():
    wb = Workbook()
    ws = wb.active
    ws.title = '历史Offer'

    headers = [
        ('Offer ID', 12), ('审批日期', 14), ('审批结果', 12),
        ('币种', 10), ('国家/城市', 14), ('职级', 10), ('职位通道', 12), ('职位类型', 12),
        ('基本月薪', 14), ('津贴及其他现金补贴', 18), ('绩效（月均）', 14), ('年终奖', 14),
        ('入职授予（年化）', 16), ('签字股票总额', 14), ('签字费总额', 14), ('安家费', 12),
        ('现金收入', 14), ('年收入', 14), ('年收入（含一次性）', 18),
        ('来源公司', 14), ('候选人现薪', 14), ('涨幅%', 10),
        ('业务紧迫度', 12), ('入职时间紧迫度', 14),
        ('能力标签', 16), ('备注', 20),
    ]
    for i, (h, w) in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    style_header(ws, 1, len(headers))

    sample = [
        'HO-2025-001', '2025-03-15', '通过',
        'CNY', '中国-北京', 'L6', '技术', '研发',
        45000, 1500, 4500, 60000, 80000, 0, 50000, 0,
        '', '', '',
        '字节跳动', 480000, 0.25,
        '正常', '1-3个月',
        '架构设计;性能优化', '',
    ]
    for i, v in enumerate(sample, 1):
        cell = ws.cell(row=2, column=i, value=v)
        cell.font = BLUE_FONT if v != '' else CELL_FONT
        cell.border = THIN_BORDER

    ws2 = wb.create_sheet('填写说明')
    notes = [
        ['字段', '说明'],
        ['Offer ID', '唯一标识，可留空由系统生成'],
        ['审批日期', '格式 YYYY-MM-DD'],
        ['审批结果', '通过 / 驳回 / 调整后通过'],
        ['来源公司', '候选人上一家公司名称，用于竞对分析'],
        ['候选人现薪', '候选人当前年薪（用于计算涨幅）'],
        ['涨幅%', '小数形式，如 0.25 = 25%'],
        ['业务紧迫度', '紧急 / 正常 / 不紧急'],
        ['入职时间紧迫度', '1个月内 / 1-3个月 / 3个月以上'],
        ['能力标签', '多个标签用分号分隔'],
    ]
    for r, row_data in enumerate(notes, 1):
        for c, v in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=v)
    style_header(ws2, 1, 2)
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 50

    wb.save(os.path.join(DIR, 'history-offer-template.xlsx'))
    print('✅ history-offer-template.xlsx')

# ========================
# 3. 规则库模板
# ========================
def gen_rule_template():
    wb = Workbook()
    ws = wb.active
    ws.title = '规则库'

    headers = [
        ('规则ID', 16), ('规则名称', 24), ('层级', 10), ('分类', 12),
        ('启用', 8), ('优先级', 10), ('条件字段', 20), ('操作符', 10), ('条件值', 16),
        ('触发类型', 10), ('触发消息', 36), ('来源', 10),
    ]
    for i, (h, w) in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    style_header(ws, 1, len(headers))

    samples = [
        ['rule_raise_cap', '涨幅上限检查（35%）', 'unified', 'raise', 'TRUE', 20, 'context.raisePercent', 'gt', '0.35', 'warn', '候选人薪资增幅超过 35%，需额外审批', 'system'],
        ['rule_raise_extreme', '涨幅红线（50%）', 'unified', 'raise', 'TRUE', 15, 'context.raisePercent', 'gt', '0.50', 'block', '候选人薪资增幅超过 50%，禁止通过', 'system'],
    ]
    for r, row_data in enumerate(samples, 2):
        for c, v in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = CELL_FONT
            cell.border = THIN_BORDER

    ws2 = wb.create_sheet('填写说明')
    notes = [
        ['字段', '说明', '可选值'],
        ['层级', '规则所属层级', 'unified / personal'],
        ['分类', '规则类别', 'bandwidth / authority / raise / fairness / structure / custom'],
        ['启用', '是否启用', 'TRUE / FALSE'],
        ['操作符', '条件判断方式', 'eq / ne / gt / gte / lt / lte / in / notIn / contains / between / regex'],
        ['触发类型', '规则命中后的动作', 'check（提示）/ warn（警告）/ block（阻断）'],
        ['来源', '规则来源', 'system / user / profile'],
    ]
    for r, row_data in enumerate(notes, 1):
        for c, v in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=v)
    style_header(ws2, 1, 3)
    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 40
    ws2.column_dimensions['C'].width = 50

    wb.save(os.path.join(DIR, 'rule-template.xlsx'))
    print('✅ rule-template.xlsx')

# ========================
# 4. 在司员工薪酬快照模板
# ========================
def gen_internal_snapshot():
    wb = Workbook()
    ws = wb.active
    ws.title = '在司员工薪酬快照'

    headers = [
        ('员工ID', 14), ('国家/城市', 14), ('职级', 10), ('职位通道', 12), ('职位类型', 12),
        ('基本月薪', 14), ('年现金收入', 14), ('年总收入', 14),
        ('入职日期', 14), ('最近调薪日', 14), ('备注', 20),
    ]
    for i, (h, w) in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    style_header(ws, 1, len(headers))

    sample = ['EMP-001', '中国-上海', 'L7', '技术', '研发', 48000, 720000, 850000, '2022-03-01', '2025-01-01', '']
    for i, v in enumerate(sample, 1):
        cell = ws.cell(row=2, column=i, value=v)
        cell.font = BLUE_FONT if v != '' else CELL_FONT
        cell.border = THIN_BORDER

    ws2 = wb.create_sheet('填写说明')
    notes = [
        ['字段', '说明'],
        ['员工ID', '脱敏后的员工标识，仅用于内部平衡参考'],
        ['国家/城市', '与 Offer 模板一致，格式：国家-城市'],
        ['职级', '当前职级'],
        ['年现金收入', '基本月薪×12+绩效+年终奖+津贴'],
        ['年总收入', '年现金收入+股票年化'],
        ['入职日期', '格式 YYYY-MM-DD'],
        ['最近调薪日', '格式 YYYY-MM-DD，用于判断薪酬时效性'],
        ['用途', '用于推荐引擎"内部平衡"因子，比对新 Offer 与在司员工水平'],
    ]
    for r, row_data in enumerate(notes, 1):
        for c, v in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=v)
    style_header(ws2, 1, 2)
    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 50

    wb.save(os.path.join(DIR, 'internal-snapshot-template.xlsx'))
    print('✅ internal-snapshot-template.xlsx')

if __name__ == '__main__':
    gen_salary_template()
    gen_history_template()
    gen_rule_template()
    gen_internal_snapshot()
    print('\n🎉 4 个 Excel 模板全部生成完成！')
